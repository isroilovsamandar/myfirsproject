from tkinter import*

import datetime
from openpyxl import Workbook, load_workbook
from random import randint



root = Tk()
root.title("Baza")
root.geometry("1200x900")
root.resizable(width=False, height=False)
root["bg"] = "light blue"

def openNewWindow():
    newWindow = Toplevel(root)
    newWindow.title("Add")
    newWindow.geometry("1200x900")
    Label(newWindow, text="Maxsulot qo'shish", font="Arial 40").pack(padx=300, pady=30)
    newWindow["bg"] = "light blue"

    name = Label(newWindow, text="Name", font=("Comic Sans MS", 35, "bold"), bg="yellow")
    name.place(x=100, y=200, width=200, height=45)
    name = Label(newWindow, text="Number", font=("Comic Sans MS", 35, "bold"), bg="yellow")
    name.place(x=100, y=280, width=200, height=45)
    name = Label(newWindow, text="Price", font=("Comic Sans MS", 35, "bold"), bg="yellow")
    name.place(x=100, y=360, width=200, height=45)
    name = Label(newWindow, text="Bar code", font=("Comic Sans MS", 35, "bold"), bg="yellow")
    name.place(x=100, y=440, width=200, height=45)
    name = Label(newWindow, text="Year", font=("Comic Sans MS", 35, "bold"), bg="yellow")
    name.place(x=100, y=520, width=200, height=45)
    name = Label(newWindow, text="Month", font=("Comic Sans MS", 35, "bold"), bg="yellow")
    name.place(x=100, y=600, width=200, height=45)
    name = Label(newWindow, text="Day", font=("Comic Sans MS", 35, "bold"), bg="yellow")
    name.place(x=100, y=680, width=200, height=45)

    e1 = Entry(newWindow, font="Arial 30")
    e1.place(x=400, y=200, width=400, height=45)
    e2 = Entry(newWindow, font="Arial 30")
    e2.place(x=400, y=280, width=400, height=45)
    e3 = Entry(newWindow, font="Arial 30")
    e3.place(x=400, y=360, width=400, height=45)
    e4 = Entry(newWindow, font="Arial 30")
    e4.place(x=400, y=440, width=400, height=45)
    e5 = Entry(newWindow, font="Arial 30")
    e5.place(x=400, y=520, width=400, height=45)
    e6 = Entry(newWindow, font="Arial 30")
    e6.place(x=400, y=600, width=400, height=45)
    e7 = Entry(newWindow, font="Arial 30")
    e7.place(x=400, y=680, width=400, height=45)

    def add():
        wb = Workbook()
        wl = load_workbook("manba.xlsx")
        wk = wl.active
        ws = wb.active
        Data = []
        for i in range(1, wk.max_row + 1):
            A = f"A{i}"
            B = f"B{i}"
            C = f"C{i}"
            D = f"D{i}"
            E = f"E{i}"
            F = f"F{i}"
            G = f"G{i}"
            Data.append([wk[A].value, wk[B].value, wk[C].value, wk[D].value, wk[E].value, wk[F].value, wk[G].value])
        foiz = (int(e3.get()) // 100) * 120
        now = datetime.datetime.now().strftime('%Y/%m/%d')
        mudat = datetime.datetime(int(e5.get()), int(e6.get()), int(e7.get())).strftime("%Y/%m/%d")
        Data.append([e1.get(), e2.get(), e3.get(), foiz, e4.get(), now, mudat])
        for k in Data:
            ws.append(k)
        wb.save("manba.xlsx")

    def cl():
        e1.delete(0, END)
        e2.delete(0, END)
        e3.delete(0, END)
        e4.delete(0, END)
        e5.delete(0, END)
        e6.delete(0, END)
        e7.delete(0, END)


    btt1 = Button(newWindow, text="OK", font="Arial 20", command=add)
    btt1.place(x=1100, y=800)
    btt2 = Button(newWindow, text="Clean", font="Arial 20", command=cl)
    btt2.place(x=900, y=800)

def openNewWindow1():
    Data = []
    newWindow = Toplevel(root)
    newWindow.title("View")
    newWindow.geometry("1200x900")
    Label(newWindow, text="Maxsulot ko'rish", font="Arial 40").pack(padx=300, pady=30)
    newWindow["bg"] = "light blue"
    def kormoq():
        wb = Workbook()
        wl = load_workbook("manba.xlsx")
        wk = wl.active
        ws = wb.active
        Data = []
        for i in range(1, wk.max_row + 1):
            A = f"A{i}"
            B = f"B{i}"
            C = f"C{i}"
            D = f"D{i}"
            E = f"E{i}"
            F = f"F{i}"
            G = f"G{i}"
            Data.append([wk[A].value, wk[B].value, wk[C].value, wk[D].value, wk[E].value, wk[F].value, wk[G].value])
    for i in Data:
        print(i)

def openNewWindow2():
    newWindow = Toplevel(root)
    newWindow.title("sell")
    newWindow.geometry("1200x900")
    Label(newWindow, text="Maxsulot sotish", font="Arial 40").pack(padx=300, pady=30)
    newWindow["bg"] = "light blue"
    name = Label(newWindow, text="Name", font=("Comic Sans MS", 35, "bold"), bg="yellow")
    name.place(x=100, y=200, width=200, height=45)
    name = Label(newWindow, text="Number", font=("Comic Sans MS", 35, "bold"), bg="yellow")
    name.place(x=100, y=280, width=200, height=45)
    e1 = Entry(newWindow, font="Arial 30")
    e1.place(x=400, y=200, width=400, height=45)
    e2 = Entry(newWindow, font="Arial 30")
    e2.place(x=400, y=280, width=400, height=45)
    def sell():
        wb = Workbook()
        wl = load_workbook("manba.xlsx")
        wk = wl.active
        ws = wb.active
        Data = []
        data = []
        data1 = []
        n = -1
        while True:
            #mahsulot_nomi = e1
            for i in range(2, wk.max_row + 1):
                A = f"A{i}"
                data.append(wk[A].value)
            for j in data:
                if e1.get() == j:
                    n = data.index(j)
            if n == -1:
                print("Bizda bunday mahsulot yoq!!!")
                continue
            else:
                break
        for i in range(2, wk.max_row + 1):
            B = f"B{i}"
            data1.append(wk[B].value)
        #sotish_soni = e2
        if int(e2.get()) <= data1[n]:
            data1[n] -= int(e2.get())
            wk.cell(row = n + 2, column = 2).value = data1[n]
            for i in range(1, wk.max_row + 1):
                A = f"A{i}"
                B = f"B{i}"
                C = f"C{i}"
                D = f"D{i}"
                E = f"E{i}"
                F = f"F{i}"
                G = f"G{i}"
                Data.append([wk[A].value, wk[B].value, wk[C].value, wk[D].value, wk[E].value, wk[F].value, wk[G].value])
            for k in Data:
                ws.append(k)
            wb.save("manba.xlsx")
        
        else:
            print("Bizda buncha mahsulot yo'q")
    def cl():
        e1.delete(0, END)
        e2.delete(0, END)


    btt1 = Button(newWindow, text="OK", font="Arial 20", command=sell)
    btt1.place(x=1100, y=800)
    btt2 = Button(newWindow, text="Clean", font="Arial 20", command=cl)
    btt2.place(x=900, y=800)

def openNewWindow3():
    newWindow = Toplevel(root)
    newWindow.title("Term")
    newWindow.geometry("1200x900")
    Label(newWindow, text="Maxsulot mudati", font="Arial 40").pack(padx=300, pady=30)
    newWindow["bg"] = "light blue"


def openNewWindow4():
    newWindow = Toplevel(root)
    newWindow.title("Report")
    newWindow.geometry("1200x900")
    Label(newWindow, text="Xisobot", font="Arial 40").pack(padx=300, pady=30)
    newWindow["bg"] = "light blue"

l1 = Label(root, text="Bo'lim tanlang", font="Arial 40")
l1.place(x=460, y=30)

btn1 = Button(root, text="Maxsulot qo'shish", font=("Comic Sans MS", 30, "bold"), command=openNewWindow)
btn1.place(x=100, y=150)

btn2 = Button(root, text="Maxsulot ko'rish", font=("Comic Sans MS", 30, "bold"), command=openNewWindow1)
btn2.place(x=100, y=270)

btn3 = Button(root, text="Maxsulot sotish", font=("Comic Sans MS", 30, "bold"), command=openNewWindow2)
btn3.place(x=100, y=390)

btn4 = Button(root, text="Maxsulot mudati", font=("Comic Sans MS", 30, "bold"), command=openNewWindow3)
btn4.place(x=100, y=510)

btn5 = Button(root, text="Xisobot", font=("Comic Sans MS", 30, "bold"), command=openNewWindow4)
btn5.place(x=100, y=630)





root.mainloop()


#2kT-v3W-ZNe-GiK









